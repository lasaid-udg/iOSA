from PyQt5.QtWidgets import QMessageBox
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Color, PatternFill
import cv2
import pandas as pd
import json
from multiprocessing import Process

from measures import *
from faceMesh import Facemesh
from database import ControllDb, MultimediaDb
from database import DatabaseThread, insert_db_thread_event
from log import Log
from WaitingSpinnerWidget import define_waiting_spinner, insert_process_event

ID_PATIENT_IDX = 0
ID_APNEA_STUDY_IDX = 0
BLANK = ""
YES = "1"
NO = "0"
DEGREE_1 = "1"
DEGREE_2 = "2"
DEGREE_3 = "3"
DEGREE_4 = "4"
DEGREE_5 = "5"
MALE = "0"
FEMALE = "1"
OTHER = "o"
MALE_EXCEL = "masculino"
FEMALE_EXCEL = "femenino"
OTHER_EXCEL = "otro"
ENCODE = "utf-8"
REPORT_EXPORTED_EVENT = "CSV file exported"
EXCEL_EXPORTED_EVENT = "Excel file exported"
CSV_CREATED_MSG_TITLE = "CSV Creado"
CSV_CREATED_MSG = "El CSV a sido creado exitosamente"
CSV_ERROR_MSG_TITLE = "Alerta!"
CSV_ERROR_MSG = "Error al generar el Excel.\nRevise que el excel esté cerrado"
PD_COLUMNS = ['Landmark', 'X', 'Y']

COLUMNS_COUNT = "A1:CH1"
COLUMNS_MEDICAL_RECORD = "A1:AL1"
COLUMNS_PDF = "AM1:BK1"
COLUMNS_IMAGE_METRICS = "BL1:BX1"
COLUMN_MEDICAL_RECORD_COLOR = "5B9BD5"
COLUMN_PDF_COLOR = "FF0000"
COLUMN_IMAGE_METRICS_COLOR = "FFE699"

EXCEL_DATA_COLUMNS = [
    "Acronimo",
    "Fecha Estudio",
    "ID Apnea",
    "Sexo",
    "Edad (Años)",
    "Frecuencia cardiaca (lpm)",
    "Presión arterial sistólica (mmhg)",
    "Presión arterial diastólica (mmhg)",
    "Frecuencia respiratoria (rpm)",
    "Saturación de oxigeno (%)",
    "Temperatura (º)",
    "Talla (m)",
    "circunferencia de cuello (cm)",
    "Peso (kg)",
    "IMC",
    "Uso de oxigeno",
    "Tabaquismo",
    "Exfumador",
    "Hipertención",
    "Presencia de ronquido",
    "Escala de Epworth",
    "ET CO2",
    "Escala de Mallampati",
    "PH",
    "pCO2",
    "pO2",
    "EB",
    "FVC Litro",
    "FVC%",
    "FEV1 Litros",
    "FEV1L%",
    "FEV1L/FVCL",
    "Ingesta de medicamentos",
    "Infarto agudo al miocardio",
    "Antecedente de accidente cerebrovascular",
    "Enfermedad vascular periferica",
    "Demencia",
    "EPOC",
    "Enfermedad Tejido Conectivo",
    "Enfermedad Epática",
    "Diabetes Millitus",
    "Hemiplejia",
    "Afección Renal",
    "Tumor solido",
    "Leucemia",
    "Linfoma",
    "VIH",
    "Escala de Charlson",
    "Indice de Apnea-Hipoapnea",
    "Promedio de respiraciones por minuto",
    "Indice de Apneas",
    "Apneas",
    "ÍAI (índice de apneas sin clasificar)",
    "Apneas indeterminadas",
    "ÍAO (índice de apneas obstructivas)",
    "Apneas obstructivas",
    "ÍAC (índice de apneas centrales)",
    "Apneas centrales",
    "ÍAM (índice de apneas mixtas)",
    "Apneas mixtas",
    "Indice de hipopneas",
    "Hipoapneas",
    "Eventos de ronquidos",
    "IDO (Índice de Desaturación de Oxigeno",
    "Saturación promedio",
    "Saturación <90%",
    "Desaturación menor",
    "Saturación <85%",
    "Saturación <80%",
    "Saturación basal",
    "Frecuencia de pulso promedio",
    "Proporcion de periodo CSR en el periodo de análisis",
    "Periodo de evaluación de flujo",
    "Ancho de cara (cm)",
    "Ángulo Ancho de Cara (°)",
    "Ancho intercantal (cm)",
    "Ancho Biocular (cm)",
    "Ancho mandibular (cm)",
    "Área triágulo maxilar (cm2)",
    "Ángulo Ancho de Mandibula (°)",
    "Ancho de Nariz (cm)",
    "Àngulo Nasion Mandibular (°)",
    "Ángulo Sub-Nasion Mandibular (°)",
    "Ángulo ANB (°)",
    "Distancia Submental (cm)",
    "Volumen Medio Fosa Craneal (cm3)"
]
PATIENT_CLINIC_DATA_COLUMNS = [
    "ID",
    "Iniciales",
    "ID Apnea",
    "Sexo",
    "Edad (Años)",
    "Frecuencia cardiaca (lpm)",
    "Presión arterial diastólica (mmhg)",
    "Presión arterial sistólica (mmhg)",
    "Frecuencia respiratoria (rpm)",
    "Saturación de oxigeno (%)",
    "Temperatura (º)",
    "Talla (m)",
    "circunferencia de cuello (cm)",
    "Peso (kg)",
    "Uso de oxigeno",
    "Tabaquismo",
    "Exfumador",
    "Presencia de ronquido",
    "Escala de Epworth",
    "ET CO2",
    "Escala de Mallampati",
    "ph",
    "pco2",
    "po2",
    "eb",
    "fvc_litros",
    "fvc",
    "fev1_litros",
    "fev1",
    "Ingesta de medicamentos",
    "Infarto agudo al miocardio",
    "Antecedente de accidente cerebrovascular",
    "Enfermedad vascular periferica",
    "Demencia",
    "EPOC",
    "Indice de Apnea-Hipoapnea",
    "Promedio de respiraciones por minuto",
    "Indice de Apneas",
    "Apneas",
    "ÍAI (índice de apneas sin clasificar)",
    "Apneas indeterminadas",
    "ÍAO (índice de apneas obstructivas)",
    "Apneas obstructivas",
    "ÍAC (índice de apneas centrales)",
    "Apneas centrales",
    "ÍAM (índice de apneas mixtas)",
    "Apneas mixtas",
    "Indice de hipopneas",
    "Hipoapneas",
    "Eventos de ronquidos",
    "IDO (Índice de Desaturación de Oxigeno",
    "Saturación promedio",
    "Saturación <90%",
    "Desaturación menor",
    "Saturación <85%",
    "Saturación <80%",
    "Saturación basal",
    "Frecuencia de pulso promedio",
    "Proporcion de periodo CSR en el periodo de análisis",
    "Periodo de evaluación de flujo",
]
MULTIMEDIA_REFERENCE_RESOURCES_COLUMNS = [
    "ID Paciente",
    "ID Estudio Apnea",
    "Nombre Imagen Frontal",
    "Nombre Imagen Lateral",
    "Nombre Video",
    "Nombre Audio",
    "Nombre OSA"
]
P1 = "p1"
P2 = "p2"
P3 = "p3"
P4 = "p4"
P5 = "p5"
P6 = "p6"
P7 = "p7"
P8 = "p8"
P9 = "p9"
P10 = "p10"
P11 = "p11"
P12 = "p12"
P13 = "p13"
P14 = "p14"
P15 = "p15"
P16 = "p16"
P17 = "p17"
P18 = "p18"
P19 = "p19"
LX = "lx"
LY = "ly"
RX = "fx"
RY = "fy"
PATIENT_IMAGE_POINTS_COLUMNS = [
    "ID Paciente",
    "ID Estudio Apnea",
    P1 + RX, P1 + RY, P2 + RX, P2 + RY, P3 + RX, P3 + RY, P4 + RX, P4 + RY,
    P5 + RX, P5 + RY, P6 + RX, P6 + RY, P7 + RX, P7 + RY, P8 + RX, P8 + RY,
    P9 + RX, P9 + RY, P10 + RX, P10 + RY, P11 + RX, P11 + RY, P12 + RX, P12 + RY,
    P13 + RX, P13 + RY, P14 + RX, P14 + RY, P15 + LX, P15 + LY, P16 + LX, P16 + LY,
    P17 + LX, P17 + LY, P18 + LX, P18 + LY, P19 + LX, P19 + LY
]

ID_INDEX = 0
FECHA_ESTUDIO_INDEX = 1
SEXO_INDEX = 4
EDAD_INDEX = 3

ID_METRICS = 2
ID_VITAL_SIGNS = 3
ID_RECORD = 4
ID_COMORBIDITY = 5
ID_DIAGNOSTIC_AIDS = 6

FRECUENCIA_CARDIACA_INDEX = 1
PRESION_ARTERIAL_SISTOLICA_INDEX = 2
PRESION_ARTERIAL_DIASTOLICA_INDEX = 3
FRECUENCIA_RESPIRATORIA_INDEX = 4
SATURACION_DE_OXIGENO_INDEX = 5
TEMPERATURA_INDEX = 6

TALLA_INDEX = 1
CIRCUNFERENCIA_DE_CUELLO_INDEX = 3
PESO_INDEX = 2

USO_DE_OXIGENO_INDEX = 1
TABAQUISMO_INDEX = 2
EXFUMADOR_INDEX = 3
PRESENCIA_DE_RONQUIDO_INDEX = 5
HIPERTENCION_INDEX = 7
INGESTA_DE_MEDICAMENTOS_INDEX = 8

INFARTO_AGUDO_AL_MIOCARDIO_INDEX = 1
ANTECEDENTE_DE_ACCIDENTE_CEREBROVASCULAR_INDEX = 2
ENFERMEDAD_VASCULAR_PERIFERICA_INDEX = 3
DEMENCIA_INDEX = 4
EPOC_INDEX = 5
ENFERMEDAD_TEJIDO_CONECTIVO_INDEX = 6
ENFERMEDAD_EPATICA_INDEX = 7
DIABETES_MILLITUS_INDEX = 8
HEMIPLEJIA_INDEX = 9
AFECCION_RENAL_INDEX = 10
TUMOR_SOLIDO_INDEX = 11
LEUCEMIA_INDEX = 12
LINFOMA_INDEX = 13
VIH_INDEX = 14

ESCALA_DE_EPWORTH_INDEX = 1
ET_CO2_INDEX = 2
ESCALA_DE_MALLAMPATI_INDEX = 3
PH_INDEX = 4
PCO2_INDEX = 5
PO2_INDEX = 6
EB_INDEX = 7
FVC_LITRO_INDEX = 8
FVC_INDEX = 9
FEV1_LITROS_INDEX = 10
FEV1L_INDEX = 11
FEV1L_FVCL_INDEX = -1

INDICE_DE_APNEA_HIPOAPNEA_INDEX = 0
INDICE_DE_APNEAS_INDEX = 1
IAI_INDEX = 2
IAO_INDEX = 3
IAC_INDEX = 4
IAM_INDEX = 5
INDICE_DE_HIPOPNEAS_INDEX = 6
IDO_INDEX = 7
SATURACION_PROMEDIO_INDEX = 8
DESATURACION_MENOR_INDEX = 9
SATURACION_BASAL_INDEX = 10
FRECUENCIA_DE_PULSO_PROMEDIO_INDEX = 11
PROMEDIO_DE_RESPIRACIONES_POR_MINUTO_INDEX = 12
PROPORCION_DE_PERIODO_CSR_EN_EL_PERIODO_DE_ANALISIS_INDEX = 12
APNEAS_INDEX = 13
APNEAS_INDETERMINADAS_INDEX = 14
APNEAS_OBSTRUCTIVAS_INDEX = 15
APNEAS_CENTRALES_INDEX = 16
APNEAS_MIXTAS_INDEX = 17
HIPOAPNEAS_INDEX = 18
EVENTOS_DE_RONQUIDOS_INDEX = 19
SATURACION_90_INDEX = 20
SATURACION_85_INDEX = 21
SATURACION_80_INDEX = 22
PERIODO_DE_EVALUACION_DE_FLUJO_INDEX = 22

IMAGE_PATH = 2
IMAGE_COORDINATES = 3
IMAGE_TAG = 4
FRONT = 0
LATERAL = 1
ANCHO_DE_CARA_INDEX = -1
ANGULO_ANCHO_DE_CARA_INDEX = -1
ANCHO_INTERCANTAL_INDEX = -1
ANCHO_BIOCULAR_INDEX = -1
ANCHO_MANDIBULAR_INDEX = -1
AREA_TRIANULO_MAXILAR_INDEX = -1
ANGULO_ANCHO_DE_CARA_INDEX = -1
ANCHO_DE_NARIZ_INDEX = -1
ANGULO_NASION_MANDIBULAR_INDEX = -1
ANGULO_SUB_NASION_MANDIBULAR_INDEX = -1
ANGULO_ANB_INDEX = -1
DISTANCIA_SUBMENTAL_INDEX = -1
VOLUMEN_MEDIO_FOSA_CRANEAL_INDEX = -1

VIDEO_PATH_INDEX = 2
AUDIO_PATH_INDEX = 2
OSA_PATH_INDEX = 2

def start_waitting_spinner() -> None:
    define_waiting_spinner()

class ExcelFile():
    def __init__(self, path):
        self.export_path = path
        self._header = EXCEL_DATA_COLUMNS

    def generate_excel(self):
        waitting_spinner_process = Process(target=start_waitting_spinner)
        waitting_spinner_process.start()
        insert_process_event(waitting_spinner_process.pid)
        # sleep(5)
        wb = Workbook()
        ws = wb.active

        self.add_header(ws)
        self.generate_list(ws)

        try:
            wb.save(self.export_path)
            log = Log()
            log.insert_log_info(EXCEL_EXPORTED_EVENT)
            waitting_spinner_process.kill()
            self.excel_exported_popup()
        except (PermissionError, FileNotFoundError):
            log = Log()
            log.insert_log_error()
            waitting_spinner_process.kill()
            self.error_excel_saving()

    def add_header(self, ws):
        ws.page_setup.fitToWidth = 0

        ws.append(self._header)

        text_format = Alignment(
            wrap_text=True, horizontal="centerContinuous", vertical="center")

        for row in ws[COLUMNS_COUNT]:
            for cell in row:
                if cell.column_letter == "A":
                    cell_color = self.change_color("A")

                if cell.column_letter == "AW":
                    cell_color = self.change_color("AW")

                if cell.column_letter == "BV":
                    cell_color = self.change_color("BV")
                
                cell.fill = cell_color

                ws.column_dimensions[cell.column_letter].width = 20
                cell.alignment = text_format

    def change_color(self, column):
        if column == "A":
            color = Color(rgb=COLUMN_MEDICAL_RECORD_COLOR)
            cell_color = PatternFill("solid", fgColor=color)

        if column == "AW":
            color = Color(rgb=COLUMN_PDF_COLOR)
            cell_color = PatternFill("solid", fgColor=color)

        if column == "BV":
            color = Color(rgb=COLUMN_IMAGE_METRICS_COLOR)
            cell_color = PatternFill("solid", fgColor=color)

        return cell_color

    def generate_list(self, ws):
        db_controll = ControllDb()
        db_multimedia = MultimediaDb()
        db_thread = DatabaseThread(
            target=db_controll.select_all_from_patients,
            args=()
        )
        db_thread.start()
        insert_db_thread_event(db_thread.native_id)
        patients = db_thread.join()

        for patient in patients:
            db_thread = DatabaseThread(
                target=db_controll.select_patient_id_date_apena_studies,
                args=(patient[ID_PATIENT_IDX],)
            )
            db_thread.start()
            insert_db_thread_event(db_thread.native_id)
            apnea_studies = db_thread.join()
            for apnea_study in apnea_studies:
                print("Apnea Study ID:", apnea_study[ID_APNEA_STUDY_IDX])
                data = []

                data.append(patient[2])
                data.append(apnea_study[ID_APNEA_STUDY_IDX])
                data.append(self.patient_sex(patient[SEXO_INDEX]))
                patient_age = self.calculate_age(patient[EDAD_INDEX],
                                               apnea_study[FECHA_ESTUDIO_INDEX])
                data.append(patient_age)

                data.insert(1, apnea_study[1].strftime("%d-%m-%Y"))

                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromMedicalRecord,
                    args=(apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                medical_record = db_thread.join()
                
                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromMetrics,
                    args=(medical_record[ID_METRICS],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                metrics  = db_thread.join()
            
                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromVitalSigns,
                    args=(medical_record[ID_VITAL_SIGNS],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                vital_signs = db_thread.join()
                
                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromHistory,
                    args=(medical_record[ID_RECORD],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                record = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromCharlsonComorbidity,
                    args=(medical_record[ID_COMORBIDITY],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                charlson_comorbidity = db_thread.join()
                
                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromAuxDiagnostic,
                    args=(medical_record[ID_DIAGNOSTIC_AIDS],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                diagnostic_aids = db_thread.join()

                data.append(str(vital_signs[FRECUENCIA_CARDIACA_INDEX]))
                data.append(str(vital_signs[PRESION_ARTERIAL_SISTOLICA_INDEX]))
                data.append(str(vital_signs[PRESION_ARTERIAL_DIASTOLICA_INDEX]))
                data.append(str(vital_signs[FRECUENCIA_RESPIRATORIA_INDEX]))
                data.append(str(vital_signs[SATURACION_DE_OXIGENO_INDEX]))
                data.append(str(float(vital_signs[TEMPERATURA_INDEX])))

                data.append(str(float(metrics[TALLA_INDEX])))
                data.append(str(float(metrics[CIRCUNFERENCIA_DE_CUELLO_INDEX])))
                data.append(str(float(metrics[PESO_INDEX])))
                data.append(str(metrics[2]/metrics[1]**2))

                data.append(self.yes_no_number(record[USO_DE_OXIGENO_INDEX]))
                data.append(self.yes_no_number(record[TABAQUISMO_INDEX]))
                data.append(self.yes_no_number(record[EXFUMADOR_INDEX]))
                data.append(self.yes_no_number(record[HIPERTENCION_INDEX]))
                data.append(self.yes_no_number(record[PRESENCIA_DE_RONQUIDO_INDEX]))

                data.append(diagnostic_aids[ESCALA_DE_EPWORTH_INDEX])
                data.append(diagnostic_aids[ET_CO2_INDEX])
                data.append(self.scale_mallampati_numner(diagnostic_aids[ESCALA_DE_MALLAMPATI_INDEX]))
                data.append(diagnostic_aids[PH_INDEX])
                data.append(diagnostic_aids[PCO2_INDEX])
                data.append(diagnostic_aids[PO2_INDEX])
                data.append(diagnostic_aids[EB_INDEX])
                data.append(diagnostic_aids[FVC_LITRO_INDEX])
                data.append(diagnostic_aids[FVC_INDEX])
                data.append(diagnostic_aids[FEV1_LITROS_INDEX])
                data.append(diagnostic_aids[FEV1L_INDEX])
                data.append(self.fev1lfvclitros(diagnostic_aids[FEV1_LITROS_INDEX],
                                                diagnostic_aids[FVC_LITRO_INDEX]))

                data.append(self.check_medicine(record[8]))

                data.append(self.yes_no_number(charlson_comorbidity[INFARTO_AGUDO_AL_MIOCARDIO_INDEX]))
                data.append(self.yes_no_number(charlson_comorbidity[ANTECEDENTE_DE_ACCIDENTE_CEREBROVASCULAR_INDEX]))
                data.append(self.yes_no_number(charlson_comorbidity[ENFERMEDAD_VASCULAR_PERIFERICA_INDEX]))
                data.append(self.yes_no_number(charlson_comorbidity[DEMENCIA_INDEX]))
                data.append(self.yes_no_number(charlson_comorbidity[EPOC_INDEX]))
                data.append(str(charlson_comorbidity[ENFERMEDAD_TEJIDO_CONECTIVO_INDEX]))
                data.append(str(charlson_comorbidity[ENFERMEDAD_EPATICA_INDEX]))
                data.append(str(charlson_comorbidity[DIABETES_MILLITUS_INDEX]))
                data.append(str(charlson_comorbidity[HEMIPLEJIA_INDEX]))
                data.append(str(charlson_comorbidity[AFECCION_RENAL_INDEX]))
                data.append(str(charlson_comorbidity[TUMOR_SOLIDO_INDEX]))
                data.append(str(charlson_comorbidity[LEUCEMIA_INDEX]))
                data.append(str(charlson_comorbidity[LINFOMA_INDEX]))
                data.append(str(charlson_comorbidity[VIH_INDEX]))

                data.append(self.__charlson_scale_sum(charlson_comorbidity, patient_age))

                db_thread = DatabaseThread(
                    target=db_controll.pdf_to_excel,
                    args=(apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                pdf = db_thread.join()

                if pdf is not None:
                    for i in pdf:
                        data.append(i)
                else:
                    for i in self.blank_fields():
                        data.append(i)

                # print(pdf)

                # print(data)

                data = self.get_measures(data, apnea_study[ID_APNEA_STUDY_IDX],
                                         db_multimedia)
                # print(data)

                ws.append(data)

    def get_measures(self, data, id_apnea_study, db_multimedia):
        db_thread = DatabaseThread(
            target=db_multimedia.get_front_photo_data,
            args=(id_apnea_study,)
        )
        db_thread.start()
        insert_db_thread_event(db_thread.native_id)
        front_photo = db_thread.join()
        db_thread = DatabaseThread(
            target=db_multimedia.get_lateral_photo_data,
            args=(id_apnea_study,)
        )
        db_thread.start()
        insert_db_thread_event(db_thread.native_id)
        lateral_photo = db_thread.join()
        # Select all the images of the same apnea study
        study_images = [front_photo, lateral_photo]
        #print("Images:", study_images)
        # Condition in case of no images loaded on that apnea study
        if study_images[0] is not None and study_images[1] is not None:
            for image_data in study_images:
                df = self.json_to_list(json.loads(
                                        json.loads(
                                        image_data[IMAGE_COORDINATES])))

                img_path = image_data[IMAGE_PATH]
                im = cv2.imread(img_path)
                facemesh = Facemesh(im, [], [])
                facemesh.qrCode()

                self.m = Measurements(df, facemesh.width, facemesh.height)
                self.m.generateScale(facemesh.qrs.qrCoords)

                if int(image_data[IMAGE_TAG]) == FRONT:
                    data = self.get_front_measurements(data)

                elif int(image_data[IMAGE_TAG]) == LATERAL:
                    data = self.get_profile_measurements(data, front_photo)

                # print("IMG SIZE:", image.shape)
                # print(df)

            return data

        return data

    def json_to_list(self, json):
        df = pd.DataFrame.from_dict(json)
        df.index.name = "Landmark"
        df.index = df.index.astype("int64")
        return df

    def get_front_measurements(self, data):
        data.append(str(self.m.getFaceWidth()))
        data.append(str(self.m.getFaceWidthAngle()))
        data.append(str(self.m.getIntercantalWidth()))
        data.append(str(self.m.getBiocularWidth()))
        data.append(str(self.m.getJawWidth()))
        data.append(str(self.m.getAreaMaxilarTriangle()))
        data.append(str(self.m.get_mandibular_width_angle()))
        data.append(str(self.m.get_nose_width()))

        return data

    def get_profile_measurements(self, data, front_photo):
        face_width = self.get_front_face_width(self.m.scale, self.m.w,
                                               front_photo)

        data.append(str(self.m.get_mandibular_nasion_angle()))
        data.append(str(self.m.get_mandibular_subnasion_angle()))
        data.append(str(self.m.get_anb_angle()))
        data.append(str(self.m.get_sm_gns()))
        data.append(str(self.m.get_middle_ranial_fossa_volume()*face_width))

        return data

    def get_front_face_width(self, scale, w, front_photo) -> float:
        if front_photo is None:
            faceWidth = -1
        else:
            # front[2] = coordinates
            points = json.loads(json.loads(front_photo[IMAGE_COORDINATES]))
            df = pd.DataFrame.from_dict(points)

            left = getXValueOfIndex(df, str(INDEX_EAR_LEFT))
            right = getXValueOfIndex(df, str(INDEX_EAR_RIGHT))
            faceWidth = ((right - left) * w) / scale

            # print("Face Width:", faceWidth, "cm")

        return round(faceWidth, 4)

    def calculate_age(self, born, study_date):
        return study_date.year - born.year - ((study_date.month,
                                               study_date.day) < (born.month, born.day))

    def patient_sex(self, num):
        if num == 1:
            sex = MALE
        elif num == 2:
            sex = FEMALE
        else:
            sex = OTHER

        return sex

    def yes_no_number(self, num):
        if num == 0:
            return NO
        else:
            return YES
        
    def __charlson_scale_sum(self, charlson_comorbidity_answers: list, patient_age: int) -> int:
        charlson_sum = 0

        if patient_age < 50:
            charlson_sum += 0
        elif 49 < patient_age < 60:
            charlson_sum += 1
        elif 59 < patient_age < 70:
            charlson_sum += 1
        elif 69 < patient_age < 80:
            charlson_sum += 2
        elif patient_age > 79:
            charlson_sum += 3

        if charlson_comorbidity_answers[HEMIPLEJIA_INDEX] == 1:
            charlson_sum += 2
        
        if charlson_comorbidity_answers[AFECCION_RENAL_INDEX] == 1:
            charlson_sum += 2
        
        if charlson_comorbidity_answers[LEUCEMIA_INDEX] == 1:
            charlson_sum += 2
        
        if charlson_comorbidity_answers[LINFOMA_INDEX] == 1:
            charlson_sum += 2
        
        if charlson_comorbidity_answers[TUMOR_SOLIDO_INDEX] == 1:
            charlson_sum += 2
        elif charlson_comorbidity_answers[TUMOR_SOLIDO_INDEX] == 2:
            charlson_sum += 3
        
        if charlson_comorbidity_answers[VIH_INDEX] == 1:
            charlson_sum += 2
        
        charlson_sum += charlson_comorbidity_answers[INFARTO_AGUDO_AL_MIOCARDIO_INDEX]
        charlson_sum += charlson_comorbidity_answers[ANTECEDENTE_DE_ACCIDENTE_CEREBROVASCULAR_INDEX]
        charlson_sum += charlson_comorbidity_answers[ENFERMEDAD_VASCULAR_PERIFERICA_INDEX]
        charlson_sum += charlson_comorbidity_answers[DEMENCIA_INDEX]
        charlson_sum += charlson_comorbidity_answers[EPOC_INDEX]
        charlson_sum += charlson_comorbidity_answers[ENFERMEDAD_TEJIDO_CONECTIVO_INDEX]
        charlson_sum += charlson_comorbidity_answers[ENFERMEDAD_EPATICA_INDEX]
        charlson_sum += charlson_comorbidity_answers[DIABETES_MILLITUS_INDEX]

        return charlson_sum

    def scale_mallampati_numner(self, num):
        if num == int(DEGREE_1):
            return DEGREE_1
        elif num == int(DEGREE_2):
            return DEGREE_2
        elif num == int(DEGREE_3):
            return DEGREE_3
        elif num == int(DEGREE_4):
            return DEGREE_4
        elif num == int(DEGREE_5):
            return DEGREE_5
        else:
            return BLANK

    def fev1lfvclitros(self, fevl, fvfl) -> float:
        try:
            FEV1litros = float(fevl)
            FVClitros = float(fvfl)
        except:
            FEV1litros = 0
            FVClitros = 1

        try:
            division = FEV1litros / FVClitros
        except (ZeroDivisionError):
            division = -1

        return division

    def check_medicine(self, text):
        if len(text) == 0:
            return NO
        else:
            return YES

    def blank_fields(self):
        blank_list = []
        for i in range(25):
            blank_list.append("")

        return blank_list

    def error_excel_saving(self):
        msg = QMessageBox()
        msg.setWindowTitle("Alerta!")
        msg.setText(
            "Error al generar el Excel.\nRevise que el excel esté cerrado")
        msg.exec_()

    def excel_exported_popup(self):
        msg = QMessageBox()
        msg.setWindowTitle("Excel Creado")
        msg.setText("El Excel a sido creado exitosamente")
        msg.exec_()


class PatientClinicData():
    def __init__(self, path) -> None:
        self._column_list = PATIENT_CLINIC_DATA_COLUMNS
        self._path = path

    def generate_report(self) -> None:
        waitting_spinner_process = Process(target=start_waitting_spinner)
        waitting_spinner_process.start()
        insert_process_event(waitting_spinner_process.pid)
        df = self.__create_dataframe()
        try:
            df.to_csv(self._path, index=False, header=True, encoding=ENCODE)
            log = Log()
            log.insert_log_info(REPORT_EXPORTED_EVENT)
            waitting_spinner_process.kill()
            self.excel_exported_popup()
        except (PermissionError, FileNotFoundError):
            log = Log()
            log.insert_log_error()
            waitting_spinner_process.kill()
            self.error_excel_saving_popup()

    def __create_dataframe(self) -> pd.DataFrame:
        all_data_rows = list()
        db_controll = ControllDb()
        db_thread = DatabaseThread(
            target=db_controll.select_all_from_patients,
            args=()
        )
        db_thread.start()
        insert_db_thread_event(db_thread.native_id)
        patients = db_thread.join()

        for patient in patients:
            
            db_thread = DatabaseThread(
                target=db_controll.select_patient_id_date_apena_studies,
                args=(patient[ID_PATIENT_IDX],)
            )
            db_thread.start()
            insert_db_thread_event(db_thread.native_id)
            apnea_studies = db_thread.join()
            for apnea_study in apnea_studies:
                data = []

                data.append(patient[0])
                data.append(patient[2])
                data.append(apnea_study[ID_APNEA_STUDY_IDX])
                data.append(self.patient_sex(patient[4]))
                # print(apnea_study[1], patient[3])
                patient_age = self.calculate_age(patient[3], apnea_study[1])
                data.append(patient_age)

                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromMedicalRecord,
                    args=(apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                medical_record = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromMetrics,
                    args=(medical_record[ID_METRICS],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                metrics  = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromVitalSigns,
                    args=(medical_record[ID_VITAL_SIGNS],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                vital_signs = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromHistory,
                    args=(medical_record[ID_RECORD],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                record = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromCharlsonComorbidity,
                    args=(medical_record[ID_COMORBIDITY],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                charlson_comorbidity = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_controll.selectAllFromAuxDiagnostic,
                    args=(medical_record[ID_DIAGNOSTIC_AIDS],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                diagnostic_aids = db_thread.join()

                data.append(str(vital_signs[1]))
                data.append(str(vital_signs[2]))
                data.append(str(vital_signs[3]))
                data.append(str(vital_signs[4]))
                data.append(str(vital_signs[5]))
                data.append(str(float(vital_signs[6])))

                data.append(str(float(metrics[1])))
                data.append(str(float(metrics[3])))
                data.append(str(float(metrics[2])))
                data.append(str(metrics[2]/metrics[1]**2))

                data.append(self.yes_no_number(record[1]))
                data.append(self.yes_no_number(record[2]))
                data.append(self.yes_no_number(record[3]))
                data.append(self.yes_no_number(record[5]))

                data.append(diagnostic_aids[1])
                data.append(diagnostic_aids[2])
                data.append(self.scale_mallampati_numner(diagnostic_aids[3]))
                data.append(diagnostic_aids[4])
                data.append(diagnostic_aids[5])
                data.append(diagnostic_aids[6])
                data.append(diagnostic_aids[7])
                data.append(diagnostic_aids[8])
                data.append(diagnostic_aids[9])
                data.append(diagnostic_aids[10])
                data.append(diagnostic_aids[11])

                data.append(self.check_medicine(record[8]))

                data.append(self.yes_no_number(charlson_comorbidity[1]))
                data.append(self.yes_no_number(charlson_comorbidity[2]))
                data.append(self.yes_no_number(charlson_comorbidity[3]))
                data.append(self.yes_no_number(charlson_comorbidity[4]))
                data.append(self.yes_no_number(charlson_comorbidity[5]))

                db_thread = DatabaseThread(
                    target=db_controll.pdf_to_excel,
                    args=(apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                pdf = db_thread.join()

                if pdf is not None:
                    for i in pdf:
                        data.append(i)
                else:
                    for i in self.blank_fields():
                        data.append(i)

                all_data_rows.append(data)

        df = pd.DataFrame(all_data_rows, columns=PATIENT_CLINIC_DATA_COLUMNS)

        return df

    def calculate_age(self, born, study_date) -> int:

        return study_date.year - born.year - ((study_date.month,
                                               study_date.day) < (born.month, born.day))

    def patient_sex(self, num) -> str:
        if num == 1:
            sex = MALE
        elif num == 2:
            sex = FEMALE
        else:
            sex = OTHER

        return sex

    def yes_no_number(self, num) -> str:
        if num == 0:
            return NO
        else:
            return YES

    def scale_mallampati_numner(self, num) -> str:
        if num == int(DEGREE_1):
            return DEGREE_1
        elif num == int(DEGREE_2):
            return DEGREE_2
        elif num == int(DEGREE_3):
            return DEGREE_3
        elif num == int(DEGREE_4):
            return DEGREE_4
        elif num == int(DEGREE_5):
            return DEGREE_5
        else:
            return BLANK

    def check_medicine(self, text) -> str:
        if len(text) == 0:
            return NO
        else:
            return YES

    def blank_fields(self) -> list:
        blank_list = []
        for i in range(25):
            blank_list.append("")

        return blank_list

    def error_excel_saving_popup(self) -> None:
        msg = QMessageBox()
        msg.setWindowTitle(CSV_ERROR_MSG_TITLE)
        msg.setText(CSV_ERROR_MSG)
        msg.exec_()

    def excel_exported_popup(self) -> None:
        msg = QMessageBox()
        msg.setWindowTitle(CSV_CREATED_MSG_TITLE)
        msg.setText(CSV_CREATED_MSG)
        msg.exec_()


class MultimediaReferenceResources():
    def __init__(self, path) -> None:
        self._column_list = MULTIMEDIA_REFERENCE_RESOURCES_COLUMNS
        self._path = path

    def generate_report(self) -> None:
        waitting_spinner_process = Process(target=start_waitting_spinner)
        waitting_spinner_process.start()
        insert_process_event(waitting_spinner_process.pid)
        df = self.__create_dataframe()
        try:
            df.to_csv(self._path, index=False, header=True, encoding=ENCODE)
            log = Log()
            log.insert_log_info(REPORT_EXPORTED_EVENT)
            waitting_spinner_process.kill()
            self.__excel_exported_popup()
        except (PermissionError, FileNotFoundError):
            log = Log()
            log.insert_log_error()
            waitting_spinner_process.kill()
            self.__error_excel_saving_popup()

    def __create_dataframe(self) -> pd.DataFrame:
        all_data_rows = list()

        db_controll = ControllDb()
        db_multimedia = MultimediaDb()
        db_thread = DatabaseThread(
            target=db_controll.select_all_from_patients,
            args=()
        )
        db_thread.start()
        insert_db_thread_event(db_thread.native_id)
        patients = db_thread.join()
        for patient in patients:
            db_thread = DatabaseThread(
                target=db_controll.select_patient_id_date_apena_studies,
                args=(patient[ID_PATIENT_IDX],)
            )
            db_thread.start()
            insert_db_thread_event(db_thread.native_id)
            apnea_studies = db_thread.join()
            for apnea_study in apnea_studies:
                data = []
                data.append(str(patient[ID_PATIENT_IDX]))
                data.append(apnea_study[ID_APNEA_STUDY_IDX])

                db_thread = DatabaseThread(
                    target=db_multimedia.get_front_photo_data,
                    args=(apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                front_photo = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_multimedia.get_lateral_photo_data,
                    args=(apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                lateral_photo = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_multimedia.get_video_data,
                    args=(apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                video = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_multimedia.get_osa_data,
                    args=(apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                osa = db_thread.join()

                front_photo_path = self.__handle_image(front_photo)
                lateral_photo_path = self.__handle_image(lateral_photo)
                data.append(front_photo_path)
                data.append(lateral_photo_path)

                video_path = self.__handle_video(video)
                data.append(video_path)

                audio_path = self.__handle_audio(video)
                data.append(audio_path)

                osa_path = self.__handle_osa(osa)
                data.append(osa_path)

                all_data_rows.append(data)

        df = pd.DataFrame(all_data_rows, columns=self._column_list)
        return df

    def __handle_image(self, photo: list) -> list:
        if photo is None:
            photo_path = BLANK
        else:
            photo_path = photo[IMAGE_PATH]

        return photo_path

    def __handle_video(self, video_tuple: tuple) -> str:
        if video_tuple is None:
            video_path = BLANK
        else:
            video_path = video_tuple[VIDEO_PATH_INDEX]

        return video_path

    def __handle_audio(self, video_tuple: tuple) -> str:
        if video_tuple is None:
            audio_path = BLANK
        else:
            lst = video_tuple[VIDEO_PATH_INDEX].split('\\')
            audio_path = r"Multimedia\Audios" + '\\' + lst[-1]
            audio_path = audio_path[:-1]
            audio_path += "3"

        return audio_path

    def __handle_osa(self, osa_tuple: tuple) -> str:
        if osa_tuple is None:
            osa_path = BLANK
        else:
            osa_path = osa_tuple[OSA_PATH_INDEX]

        return osa_path

    def __error_excel_saving_popup(self) -> None:
        msg = QMessageBox()
        msg.setWindowTitle(CSV_ERROR_MSG_TITLE)
        msg.setText(CSV_ERROR_MSG)
        msg.exec_()

    def __excel_exported_popup(self) -> None:
        msg = QMessageBox()
        msg.setWindowTitle(CSV_CREATED_MSG_TITLE)
        msg.setText(CSV_CREATED_MSG)
        msg.exec_()


class PatientImagePointsReport():
    def __init__(self, path) -> None:
        self._column_list = PATIENT_IMAGE_POINTS_COLUMNS
        self._path = path

    def generate_report(self) -> None:
        waitting_spinner_process = Process(target=start_waitting_spinner)
        waitting_spinner_process.start()
        insert_process_event(waitting_spinner_process.pid)
        df = self.__create_dataframe()
        try:
            df.to_csv(self._path, index=False, header=True, encoding=ENCODE)
            log = Log()
            log.insert_log_info(REPORT_EXPORTED_EVENT)
            waitting_spinner_process.kill()
            self.__excel_exported_popup()
        except (PermissionError, FileNotFoundError):
            log = Log()
            log.insert_log_error()
            waitting_spinner_process.kill()
            self.__error_excel_saving_popup()

    def __create_dataframe(self) -> pd.DataFrame:
        all_data_rows = list()
        db_controll = ControllDb()
        db_multimedia = MultimediaDb()
        db_thread = DatabaseThread(
            target=db_controll.select_all_id_patients,
            args=()
        )
        db_thread.start()
        insert_db_thread_event(db_thread.native_id)
        id_patients = db_thread.join()
        for id_patient in id_patients:
            db_thread = DatabaseThread(
                target=db_controll.select_all_id_apnea_studies,
                args=(id_patient[ID_PATIENT_IDX],)
            )
            db_thread.start()
            insert_db_thread_event(db_thread.native_id)
            id_apnea_studies = db_thread.join()
            for id_apnea_study in id_apnea_studies:
                data = []

                data.append(id_patient[ID_PATIENT_IDX])
                data.append(id_apnea_study[ID_APNEA_STUDY_IDX])

                db_thread = DatabaseThread(
                    target=db_multimedia.get_front_photo_data,
                    args=(id_apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                front_photo = db_thread.join()

                db_thread = DatabaseThread(
                    target=db_multimedia.get_lateral_photo_data,
                    args=(id_apnea_study[ID_APNEA_STUDY_IDX],)
                )
                db_thread.start()
                insert_db_thread_event(db_thread.native_id)
                lateral_photo = db_thread.join()

                coords_db = self.__handle_image(front_photo, lateral_photo)
                data.extend(self.__handle_coords(coords_db))

                all_data_rows.append(data)

        df = pd.DataFrame(all_data_rows, columns=self._column_list)

        return df
    
    def __handle_image(self, front_photo: tuple, lateral_photo: tuple) -> list:
        coordinates = list()

        if front_photo is not None:
            coordinates.append(front_photo[IMAGE_COORDINATES])
        else:
            pass
        if lateral_photo is not None:
            coordinates.append(lateral_photo[IMAGE_COORDINATES])
        else:
            pass

        return coordinates

    def __handle_coords(self, coords_db: tuple) -> list:
        coords = list()
        if len(coords_db) == 0:
            coords = self.__no_coords(coords, 0)
            coords = self.__no_coords(coords, 1)
        else:
            for tuple_coords in coords_db:
                coords_dict = json.loads(json.loads(tuple_coords))
                zip_coords = list(zip(coords_dict["X"].values(),
                                      coords_dict["Y"].values()))

                for coord_tuple in zip_coords:
                    coords.append(coord_tuple[0])
                    coords.append(coord_tuple[1])

                if len(coords_db) < 2:
                    coords = self.__no_coords(coords, int(tuple_coords[1]))

        return coords

    def __no_coords(self, lst: list, tag: int) -> list:
        if tag == 1:
            for i in range(28):
                lst.insert(0, BLANK)
        elif tag == 0:
            for i in range(10):
                lst.append(BLANK)

        return lst

    def __error_excel_saving_popup(self) -> None:
        msg = QMessageBox()
        msg.setWindowTitle(CSV_ERROR_MSG_TITLE)
        msg.setText(CSV_ERROR_MSG)
        msg.exec_()

    def __excel_exported_popup(self) -> None:
        msg = QMessageBox()
        msg.setWindowTitle(CSV_CREATED_MSG_TITLE)
        msg.setText(CSV_CREATED_MSG)
        msg.exec_()
